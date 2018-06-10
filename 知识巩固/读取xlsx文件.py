from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    #打开文件
    file=load_workbook(filename=path)
    #获取所有的表格名
    # sheets=file.get_sheet_names()
    #
    # sheet=file.get_sheet_by_name(sheets[0])
    # sheets = file.get_sheet_names()
    sheets=file.sheetnames
    # 拿出一个表格
    sheet = file.get_sheet_by_name(sheets[0])
    for lineNum in range(1,sheet.max_row+1):
        lineList=[]
        for columNum in range(1,sheet.max_column+1):
            data=sheet.cell(row=lineNum,column=columNum).value
            lineList.append(data)
        print(lineList)


path=r'C:\Users\User\Desktop\杂类文件\2016-2017学年国家励志奖学金汇总表.xlsx'
readXlsxFile(path)